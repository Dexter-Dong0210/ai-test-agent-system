#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test Case Excel Export Script
Converts structured test case JSON into a formatted Excel workbook.
"""

import json
import sys
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def set_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def set_cell_border(cell):
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def auto_width(worksheet, min_width=10, max_width=60):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        adjusted_width = min(max(length + 2, min_width), max_width)
        worksheet.column_dimensions[col_letter].width = adjusted_width


def format_steps(steps):
    """将步骤列表格式化为 step1:xxx\nstep2:xxx 格式"""
    if not steps:
        return ""
    if isinstance(steps, str):
        # 尝试按行分割
        lines = [line.strip() for line in steps.split("\n") if line.strip()]
    else:
        lines = [str(s).strip() for s in steps if s]
    
    formatted = []
    for idx, line in enumerate(lines, 1):
        # 移除原有的数字前缀（如 "1."、"1、"等）
        line_clean = re.sub(r'^(\d+[\.、]\s*)', '', line)
        formatted.append(f"step{idx}:{line_clean}")
    return "\n".join(formatted)


def format_assertions(results):
    """将预期结果列表格式化为 assert1:xxx\nassert2:xxx 格式"""
    if not results:
        return ""
    if isinstance(results, str):
        lines = [line.strip() for line in results.split("\n") if line.strip()]
    else:
        lines = [str(r).strip() for r in results if r]
    
    formatted = []
    for idx, line in enumerate(lines, 1):
        # 移除原有的数字前缀
        line_clean = re.sub(r'^(\d+[\.、]\s*)', '', line)
        formatted.append(f"assert{idx}:{line_clean}")
    return "\n".join(formatted)


def create_cases_sheet(wb, cases):
    ws = wb.active
    ws.title = "用例详情"

    headers = [
        "用例ID", "模块", "标题", "优先级", "用例类型",
        "来源需求", "来源测试点", "前置条件", "测试数据",
        "测试步骤", "预期结果", "清理步骤", "备注"
    ]

    ws.append(headers)
    for cell in ws[1]:
        set_header_style(cell)

    for case in cases:
        row = [
            case.get("case_id", ""),
            case.get("module", ""),
            case.get("title", ""),
            case.get("priority", ""),
            case.get("type", ""),
            ", ".join(case.get("source_req", [])) if isinstance(case.get("source_req"), list) else case.get("source_req", ""),
            case.get("source_tp", ""),
            "\n".join(case.get("preconditions", [])) if isinstance(case.get("preconditions"), list) else case.get("preconditions", ""),
            json.dumps(case.get("test_data", {}), ensure_ascii=False) if isinstance(case.get("test_data"), dict) else str(case.get("test_data", "")),
            format_steps(case.get("steps", [])),
            format_assertions(case.get("expected_results", [])),
            "\n".join(case.get("cleanup_steps", [])) if isinstance(case.get("cleanup_steps"), list) else case.get("cleanup_steps", ""),
            case.get("remark", ""),
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            set_cell_border(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_traceability_sheet(wb, traceability):
    ws = wb.create_sheet(title="需求追踪")
    headers = ["需求ID", "需求描述", "覆盖用例数", "覆盖用例ID列表", "覆盖状态"]
    ws.append(headers)
    for cell in ws[1]:
        set_header_style(cell)

    for item in traceability:
        case_ids = item.get("case_ids", [])
        row = [
            item.get("req_id", ""),
            item.get("req_desc", ""),
            len(case_ids),
            ", ".join(case_ids),
            item.get("coverage_status", "已覆盖" if case_ids else "未覆盖"),
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            set_cell_border(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_statistics_sheet(wb, statistics, metadata):
    ws = wb.create_sheet(title="统计概览")

    # Metadata section
    ws.append(["元信息"])
    ws["A1"].font = Font(bold=True, size=12)
    meta_items = [
        ["功能名称", metadata.get("feature_name", "")],
        ["版本", metadata.get("version", "")],
        ["交付日期", metadata.get("delivery_date", "")],
        ["评审结论", metadata.get("review_status", "")],
        ["质量总分", metadata.get("quality_score", "")],
    ]
    for item in meta_items:
        ws.append(item)
    ws.append([])

    # Statistics section
    ws.append(["用例统计"])
    ws["A" + str(ws.max_row)].font = Font(bold=True, size=12)
    stat_headers = ["统计项", "数量"]
    ws.append(stat_headers)
    for cell in ws[ws.max_row]:
        set_header_style(cell)

    stat_mapping = [
        ("总用例数", "total"),
        ("P0用例", "p0"),
        ("P1用例", "p1"),
        ("P2用例", "p2"),
        ("正向用例", "positive"),
        ("反向用例", "negative"),
        ("边界用例", "boundary"),
        ("异常用例", "exception"),
    ]
    for label, key in stat_mapping:
        ws.append([label, statistics.get(key, 0)])
        for cell in ws[ws.max_row]:
            set_cell_border(cell)

    auto_width(ws)


def create_review_sheet(wb, review_data):
    ws = wb.create_sheet(title="评审结果")

    # Score summary
    ws.append(["质量评分"])
    ws["A1"].font = Font(bold=True, size=12)
    score_headers = ["评分维度", "满分", "得分", "权重", "加权得分", "说明"]
    ws.append(score_headers)
    for cell in ws[ws.max_row]:
        set_header_style(cell)

    scores = review_data.get("scores", [])
    for s in scores:
        ws.append([
            s.get("dimension", ""),
            s.get("full_score", 100),
            s.get("score", 0),
            s.get("weight", "0%"),
            s.get("weighted_score", 0),
            s.get("comment", ""),
        ])
        for cell in ws[ws.max_row]:
            set_cell_border(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.append([])
    ws.append(["问题清单"])
    ws["A" + str(ws.max_row)].font = Font(bold=True, size=12)
    issue_headers = ["序号", "用例ID", "问题类型", "问题描述", "严重程度", "修改建议", "责任人"]
    ws.append(issue_headers)
    for cell in ws[ws.max_row]:
        set_header_style(cell)

    issues = review_data.get("issues", [])
    for idx, issue in enumerate(issues, 1):
        ws.append([
            idx,
            issue.get("case_id", ""),
            issue.get("issue_type", ""),
            issue.get("description", ""),
            issue.get("severity", ""),
            issue.get("suggestion", ""),
            issue.get("owner", ""),
        ])
        for cell in ws[ws.max_row]:
            set_cell_border(cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws)
    ws.freeze_panes = "A2"


def main():
    if len(sys.argv) < 3:
        print("Usage: python export_to_excel.py <input_json_path> <output_xlsx_path>")
        sys.exit(1)

    input_json = sys.argv[1]
    output_xlsx = sys.argv[2]

    if not os.path.exists(input_json):
        print(f"Error: Input file not found: {input_json}")
        sys.exit(1)

    data = load_json(input_json)
    metadata = data.get("metadata", {})
    statistics = data.get("statistics", {})
    traceability = data.get("traceability", [])
    cases = data.get("cases", [])
    review_data = data.get("review", {})

    wb = Workbook()
    create_cases_sheet(wb, cases)
    create_traceability_sheet(wb, traceability)
    create_statistics_sheet(wb, statistics, metadata)
    create_review_sheet(wb, review_data)

    wb.save(output_xlsx)
    print(f"Excel exported successfully: {output_xlsx}")


if __name__ == "__main__":
    main()
