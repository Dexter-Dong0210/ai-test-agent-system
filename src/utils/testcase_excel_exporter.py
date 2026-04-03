"""
测试用例 Excel 导出器

支持将结构化测试用例数据导出为格式化的 Excel 文件，
符合 output-formatter Skill 中定义的样式规范。
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 样式常量
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CELL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Excel 列定义: (字段名, 列标题, 默认宽度)
EXCEL_COLUMNS = [
    ("id", "用例编号", 18),
    ("title", "用例标题", 30),
    ("preconditions", "前置条件", 30),
    ("steps", "测试步骤", 40),
    ("expected_results", "预期结果", 40),
    ("priority", "优先级", 10),
    ("module", "所属模块", 15),
    ("type", "用例类型", 12),
    ("test_data", "测试数据", 30),
    ("remarks", "备注", 20),
]


import re


_NUMBER_PREFIX_RE = re.compile(r"^[\d]+[\.、]\s*")


def _normalize_steps(steps: Union[str, List[Any], None]) -> str:
    """将步骤数据规范化为 stepN:xxx 格式文本"""
    if steps is None:
        return ""
    if isinstance(steps, str):
        lines = [line.strip() for line in steps.strip().split("\n") if line.strip()]
        formatted = []
        for i, line in enumerate(lines, 1):
            cleaned = _NUMBER_PREFIX_RE.sub("", line)
            formatted.append(f"step{i}: {cleaned}")
        return "\n".join(formatted)
    if isinstance(steps, list):
        lines = []
        for i, step in enumerate(steps, 1):
            if isinstance(step, dict):
                action = step.get("action", "")
                target = step.get("target", "")
                data = step.get("data", "")
                parts = [f"step{i}:{action}"]
                if target:
                    parts.append(f" [{target}]")
                if data:
                    parts.append(f" (数据: {data})")
                lines.append("".join(parts))
            elif isinstance(step, str):
                cleaned = _NUMBER_PREFIX_RE.sub("", step.strip())
                lines.append(f"step{i}: {cleaned}")
            else:
                lines.append(str(step))
        return "\n".join(lines)
    return str(steps)


def _normalize_test_data(test_data: Union[str, Dict[str, Any], None]) -> str:
    """将测试数据规范化为文本"""
    if test_data is None:
        return ""
    if isinstance(test_data, str):
        return test_data.strip()
    if isinstance(test_data, dict):
        return "\n".join([f"{k}: {v}" for k, v in test_data.items()])
    return str(test_data)


def _normalize_list_or_str(value: Union[str, List[Any], None]) -> str:
    """将列表或字符串规范化为文本"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        return "\n".join([f"• {v}" for v in value])
    return str(value)


def _normalize_expected_results(value: Union[str, List[Any], None]) -> str:
    """将预期结果规范化为 assertN:xxx 格式文本"""
    if value is None:
        return ""
    if isinstance(value, str):
        lines = [line.strip() for line in value.strip().split("\n") if line.strip()]
        formatted = []
        for i, line in enumerate(lines, 1):
            cleaned = _NUMBER_PREFIX_RE.sub("", line)
            formatted.append(f"assert{i}: {cleaned}")
        return "\n".join(formatted)
    if isinstance(value, list):
        formatted = []
        for i, v in enumerate(value, 1):
            cleaned = _NUMBER_PREFIX_RE.sub("", str(v))
            formatted.append(f"assert{i}: {cleaned}")
        return "\n".join(formatted)
    return str(value)


def _extract_value(case: Dict[str, Any], key: str) -> str:
    """从用例字典中提取并规范化字段值"""
    value = case.get(key, "")
    if key == "steps":
        return _normalize_steps(value)
    if key == "test_data":
        return _normalize_test_data(value)
    if key in ("preconditions", "remarks"):
        return _normalize_list_or_str(value)
    if key == "expected_results":
        return _normalize_expected_results(value)
    return str(value) if value is not None else ""


def export_testcases_to_excel(
    testcases: List[Dict[str, Any]], output_path: Union[str, Path]
) -> str:
    """
    将测试用例列表导出为 Excel 文件

    Args:
        testcases: 测试用例字典列表
        output_path: 输出文件路径

    Returns:
        导出成功的提示信息，包含文件路径
    """
    if not testcases:
        raise ValueError("测试用例列表不能为空")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "用例详情"

    # 写入表头
    for col_idx, (field, header, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写入数据
    for row_idx, case in enumerate(testcases, 2):
        for col_idx, (field, header, width) in enumerate(EXCEL_COLUMNS, 1):
            value = _extract_value(case, field)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = CELL_BORDER
        # 设置默认行高（约 60 磅对应 openpyxl 的 90）
        ws.row_dimensions[row_idx].height = 90

    # 冻结首行
    ws.freeze_panes = "A2"

    # 启用自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(str(output_path))
    logger.info(f"已导出 {len(testcases)} 条测试用例到 {output_path}")
    return f"成功导出 {len(testcases)} 条测试用例，文件路径: {output_path.resolve()}"
